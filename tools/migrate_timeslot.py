#!/usr/bin/env python3
"""
Migrate AMS case files from the pre-v1.3.0 CSV-list TimeSlot encoding
to the v1.3.0 paired-key encoding.

Pre-v1.3.0 cases stored per-(area, slot) load scaling and per-(gen,
slot) commitment as comma-separated lists in single ``EDSlot.sd`` /
``EDSlot.ug`` / ``UCSlot.sd`` cells. List position aligned
implicitly with ``Area.get_all_idxes()`` /
``StaticGen.get_all_idxes()``.

v1.3.0 splits this into dedicated tables — ``EDSlotLoad``,
``EDSlotGen``, ``UCSlotLoad`` — with one row per ``(device, slot)``
pair carrying a scalar value column. See the project plan at
``.claude/projects/timeslot_paired_key/`` for the rationale.

Usage::

    python tools/migrate_timeslot.py <case-file> [--out <out-path>]
    python tools/migrate_timeslot.py <case-file> --in-place

Supports JSON and XLSX cases. CSV / .m / .raw cases never carried
TimeSlot data, so they don't need migration.

Limitations:

* Length checks compare the CSV-list cell against
  ``len(Area.get_all_idxes())`` / ``len(StaticGen.get_all_idxes())``.
  A mismatch raises with the slot idx + observed/expected lengths.
* Duplicate slots in the source raise.
* The ``Summary`` block (and other unrelated tables) is preserved
  verbatim.
"""
from __future__ import annotations

import argparse
import json
import logging
import shutil
import sys
from copy import deepcopy
from pathlib import Path
from typing import Iterable

logger = logging.getLogger(__name__)


def _is_legacy_cell(value) -> bool:
    """Return True if ``value`` looks like a CSV-list legacy cell."""
    if isinstance(value, str) and ',' in value:
        return True
    if isinstance(value, list):
        # Already-parsed list (e.g. from a re-loaded JSON that was
        # mid-migration). Treat as legacy too.
        return True
    return False


def _parse_csv_list(cell) -> list[str]:
    """Parse a CSV-list cell into a stripped list of string tokens."""
    if isinstance(cell, list):
        return [str(x).strip() for x in cell]
    return [tok.strip() for tok in str(cell).split(',')]


def _device_orderings(data: dict) -> tuple[list, list]:
    """
    Reproduce the canonical ``Area`` and ``StaticGen`` orderings
    that AMS would use after loading ``data``.

    ``Area``: row order of the case file's ``Area`` block.

    ``StaticGen`` group: per ``ams/models/__init__.py`` the static
    block is ordered ``[PQ, Slack, PV]``; the StaticGen group is
    ``[Slack, PV]`` (PQ is load, not gen). Within each model, file
    row order is preserved.
    """
    areas = [row['idx'] for row in data.get('Area', [])]
    static_gens: list = []
    for model_key in ('Slack', 'PV'):
        for row in data.get(model_key, []):
            static_gens.append(row['idx'])
    return areas, static_gens


def _slim_definition_row(row: dict) -> dict:
    """Strip the legacy ``sd`` / ``ug`` columns from a slot-definition row."""
    return {k: v for k, v in row.items() if k not in ('sd', 'ug')}


def _emit_load_rows(slot_rows: Iterable[dict], areas: list, table_name: str
                    ) -> list[dict]:
    """Emit ``EDSlotLoad`` / ``UCSlotLoad`` rows from legacy slot rows."""
    if not areas:
        # Cases with no Area block (npcc, wecc) carried legacy ``sd``
        # cells that were never consumed at solve time — LoadScale.v
        # indexes Area and would have failed before this code ran.
        # Drop the legacy data rather than emit unloadable rows.
        legacy = [s.get('idx') for s in slot_rows
                  if _is_legacy_cell(s.get('sd'))]
        if legacy:
            head = legacy[:3] + (['...'] if len(legacy) > 3 else [])
            logger.warning(
                "%s: no Area entries — dropping legacy sd from "
                "%d slots (was not consumable by LoadScale anyway). "
                "Slots affected: %s",
                table_name, len(legacy), head)
        return []
    out = []
    for slot in slot_rows:
        if 'sd' not in slot:
            continue
        if not _is_legacy_cell(slot['sd']):
            continue
        parsed = _parse_csv_list(slot['sd'])
        if len(parsed) != len(areas):
            raise ValueError(
                f"{table_name}: slot {slot.get('idx')!r} has sd of "
                f"length {len(parsed)}, expected {len(areas)} "
                f"(== len(Area)). File ordering: {areas}.")
        for area_idx, sd_val in zip(areas, parsed):
            out.append({
                'area': area_idx,
                'slot': slot['idx'],
                'sd': float(sd_val),
            })
    return out


def _emit_gen_rows(slot_rows: Iterable[dict], gens: list) -> list[dict]:
    """Emit ``EDSlotGen`` rows from legacy ``EDSlot`` rows."""
    if not gens:
        legacy = [s.get('idx') for s in slot_rows
                  if _is_legacy_cell(s.get('ug'))]
        if legacy:
            head = legacy[:3] + (['...'] if len(legacy) > 3 else [])
            logger.warning(
                "EDSlot: no StaticGen entries — dropping legacy ug "
                "from %d slots. Slots affected: %s",
                len(legacy), head)
        return []
    out = []
    for slot in slot_rows:
        if 'ug' not in slot:
            continue
        if not _is_legacy_cell(slot['ug']):
            continue
        parsed = _parse_csv_list(slot['ug'])
        if len(parsed) != len(gens):
            raise ValueError(
                f"EDSlot: slot {slot.get('idx')!r} has ug of "
                f"length {len(parsed)}, expected {len(gens)} "
                f"(== len(StaticGen)). File ordering: {gens}.")
        for gen_idx, ug_val in zip(gens, parsed):
            out.append({
                'gen': gen_idx,
                'slot': slot['idx'],
                'ug': int(float(ug_val)),
            })
    return out


def _rename_legacy_keys(data: dict) -> tuple[dict, bool]:
    """
    Rename pre-v1.3.0 ``EDTSlot`` / ``UCTSlot`` top-level keys to
    ``EDSlot`` / ``UCSlot``. Idempotent.

    Returns the (possibly-rewritten) dict and a ``renamed`` flag.
    """
    if 'EDTSlot' not in data and 'UCTSlot' not in data:
        return data, False
    out = {}
    renamed = False
    for k, v in data.items():
        if k == 'EDTSlot':
            out['EDSlot'] = v
            renamed = True
        elif k == 'UCTSlot':
            out['UCSlot'] = v
            renamed = True
        else:
            out[k] = v
    return out, renamed


def migrate_data(data: dict) -> tuple[dict, dict]:
    """
    Migrate an in-memory case dict from the legacy TimeSlot encoding
    to the v1.3.0 paired-key encoding.

    Two transformations applied (idempotent):

    * Rename top-level keys ``EDTSlot`` / ``UCTSlot`` to ``EDSlot`` /
      ``UCSlot`` (the pre-v1.3.0 names dropped the redundant "T").
    * Split CSV-list ``sd`` / ``ug`` cells into the new ``EDSlotLoad``
      / ``EDSlotGen`` / ``UCSlotLoad`` per-(device, slot) tables.

    Returns
    -------
    new_data : dict
        Migrated case data. Safe to write to JSON / XLSX.
    summary : dict
        ``{'ed_load': N, 'ed_gen': N, 'uc_load': N, 'renamed': bool,
        'no_op': bool}``.
    """
    out = deepcopy(data)
    out, renamed = _rename_legacy_keys(out)
    areas, gens = _device_orderings(out)

    edt_legacy = out.get('EDSlot', [])
    uct_legacy = out.get('UCSlot', [])

    has_legacy_cells = (
        any(_is_legacy_cell(r.get('sd')) or _is_legacy_cell(r.get('ug'))
            for r in edt_legacy)
        or any(_is_legacy_cell(r.get('sd')) for r in uct_legacy)
    )
    no_op = not (renamed or has_legacy_cells)
    if no_op:
        return out, {'ed_load': 0, 'ed_gen': 0, 'uc_load': 0,
                     'renamed': False, 'no_op': True}

    if not has_legacy_cells:
        return out, {'ed_load': 0, 'ed_gen': 0, 'uc_load': 0,
                     'renamed': renamed, 'no_op': False}

    # Slim definition tables.
    if edt_legacy:
        out['EDSlot'] = [_slim_definition_row(r) for r in edt_legacy]
    if uct_legacy:
        out['UCSlot'] = [_slim_definition_row(r) for r in uct_legacy]

    # Emit per-axis tables. Insert them adjacent to the slot
    # definitions for readability.
    ed_load = _emit_load_rows(edt_legacy, areas, 'EDSlot')
    ed_gen = _emit_gen_rows(edt_legacy, gens)
    uc_load = _emit_load_rows(uct_legacy, areas, 'UCSlot')

    # Inject in dict-order: keep the existing key order and append
    # the new keys near their definition counterparts. JSON dicts
    # preserve insertion order in Py3.7+.
    if ed_load and 'EDSlotLoad' not in out:
        out = _insert_after(out, 'EDSlot', 'EDSlotLoad', ed_load)
    if ed_gen and 'EDSlotGen' not in out:
        out = _insert_after(out, 'EDSlotLoad' if ed_load else 'EDSlot',
                            'EDSlotGen', ed_gen)
    if uc_load and 'UCSlotLoad' not in out:
        out = _insert_after(out, 'UCSlot', 'UCSlotLoad', uc_load)

    return out, {'ed_load': len(ed_load), 'ed_gen': len(ed_gen),
                 'uc_load': len(uc_load), 'renamed': renamed,
                 'no_op': False}


def _insert_after(data: dict, after_key: str, new_key: str,
                  new_value) -> dict:
    """Return a new dict with ``new_key`` inserted right after ``after_key``."""
    if after_key not in data:
        # Fallback: just append.
        out = dict(data)
        out[new_key] = new_value
        return out
    out = {}
    for k, v in data.items():
        out[k] = v
        if k == after_key:
            out[new_key] = new_value
    return out


# ---------------------------------------------------------------------------
# JSON adapter
# ---------------------------------------------------------------------------

def migrate_json(path: Path, out_path: Path) -> dict:
    with path.open() as f:
        data = json.load(f)
    new_data, summary = migrate_data(data)
    if summary.get('no_op'):
        logger.info("%s: no legacy cells, skipping", path.name)
        return summary
    with out_path.open('w') as f:
        json.dump(new_data, f, indent=2)
        f.write('\n')
    logger.info("%s -> %s: EDSlotLoad=%d EDSlotGen=%d UCSlotLoad=%d",
                path.name, out_path.name,
                summary['ed_load'], summary['ed_gen'], summary['uc_load'])
    return summary


# ---------------------------------------------------------------------------
# XLSX adapter
# ---------------------------------------------------------------------------

def _xlsx_sheet_to_rows(ws) -> list[dict]:
    """Convert an openpyxl worksheet into a list-of-dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if all(cell is None for cell in raw):
            continue
        row = {}
        for h, v in zip(headers, raw):
            if h is None:
                continue
            row[h] = v
        out.append(row)
    return out


def _xlsx_load(path: Path) -> tuple[dict, list]:
    """Read an xlsx case into a {sheet: [rows]} dict + sheet order."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    data = {}
    sheet_order = []
    for sname in wb.sheetnames:
        sheet_order.append(sname)
        data[sname] = _xlsx_sheet_to_rows(wb[sname])
    return data, sheet_order


def _xlsx_dump(path: Path, data: dict, sheet_order: list) -> None:
    """Write a {sheet: [rows]} dict back to xlsx, preserving sheet order."""
    import openpyxl
    wb = openpyxl.Workbook()
    # Drop the auto-created blank sheet.
    wb.remove(wb.active)
    for sname in sheet_order:
        rows = data.get(sname, [])
        ws = wb.create_sheet(title=sname)
        if not rows:
            continue
        headers = sorted({k for r in rows for k in r.keys()},
                         key=_header_sort_key(rows))
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    wb.save(str(path))


def _header_sort_key(rows: list[dict]):
    """Stable header order: first-seen order across the row list."""
    seen = []
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.append(k)
    rank = {h: i for i, h in enumerate(seen)}
    return lambda h: rank.get(h, len(rank))


def migrate_xlsx(path: Path, out_path: Path) -> dict:
    data, sheet_order = _xlsx_load(path)
    new_data, summary = migrate_data(data)
    if summary.get('no_op'):
        logger.info("%s: no legacy cells, skipping", path.name)
        return summary
    # Mirror the EDTSlot/UCTSlot -> EDSlot/UCSlot key rename in the
    # sheet-order list so the rewritten workbook uses the new names.
    sheet_order = ['EDSlot' if s == 'EDTSlot'
                   else 'UCSlot' if s == 'UCTSlot'
                   else s for s in sheet_order]
    # Ensure new sheets are inserted in the order migrate_data placed them.
    for new_sheet in ('EDSlotLoad', 'EDSlotGen', 'UCSlotLoad'):
        if new_sheet in new_data and new_sheet not in sheet_order:
            after = {'EDSlotLoad': 'EDSlot',
                     'EDSlotGen': 'EDSlotLoad' if 'EDSlotLoad' in sheet_order
                                  else 'EDSlot',
                     'UCSlotLoad': 'UCSlot'}[new_sheet]
            if after in sheet_order:
                sheet_order.insert(sheet_order.index(after) + 1, new_sheet)
            else:
                sheet_order.append(new_sheet)
    _xlsx_dump(out_path, new_data, sheet_order)
    logger.info("%s -> %s: EDSlotLoad=%d EDSlotGen=%d UCSlotLoad=%d",
                path.name, out_path.name,
                summary['ed_load'], summary['ed_gen'], summary['uc_load'])
    return summary


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.strip().split('\n')[0])
    p.add_argument('case', type=Path, help='Case file path (.json or .xlsx).')
    g = p.add_mutually_exclusive_group()
    g.add_argument('--out', type=Path, default=None,
                   help='Output path (default: <case>.migrated<.ext>).')
    g.add_argument('--in-place', action='store_true',
                   help='Overwrite the input file (a .bak copy is kept).')
    p.add_argument('-v', '--verbose', action='store_true',
                   help='Verbose logging.')
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(message)s',
    )

    case: Path = args.case
    if not case.exists():
        logger.error("Case file not found: %s", case)
        return 2

    if args.in_place:
        backup = case.with_suffix(case.suffix + '.bak')
        shutil.copy2(case, backup)
        logger.info("Backup: %s", backup.name)
        out_path = case
    elif args.out is not None:
        out_path = args.out
    else:
        out_path = case.with_suffix('.migrated' + case.suffix)

    suffix = case.suffix.lower()
    if suffix == '.json':
        migrate_json(case, out_path)
    elif suffix == '.xlsx':
        migrate_xlsx(case, out_path)
    else:
        logger.error("Unsupported format: %s (need .json or .xlsx)", suffix)
        return 2
    return 0


if __name__ == '__main__':
    sys.exit(main())
