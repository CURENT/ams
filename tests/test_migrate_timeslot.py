"""
Tests for ``tools/migrate_timeslot.py`` — the v1.3.0 case-file
migrator that splits legacy ``EDTSlot`` / ``UCTSlot`` CSV-list
cells into the new per-axis tables and renames the top-level keys.
"""

import json
import sys
import tempfile
import unittest
from copy import deepcopy
from pathlib import Path

import pytest

# tools/ isn't on the install path; add it for the test session.
_TOOLS = Path(__file__).resolve().parent.parent / 'tools'
sys.path.insert(0, str(_TOOLS))

from migrate_timeslot import (  # noqa: E402
    migrate_data, migrate_json, migrate_xlsx, _rename_legacy_keys,
)


def _legacy_case():
    """Minimal legacy case dict with both EDTSlot and UCTSlot."""
    return {
        'Area': [{'idx': 1, 'u': 1.0, 'name': '1'},
                 {'idx': 2, 'u': 1.0, 'name': '2'}],
        'Slack': [{'idx': 'S1', 'u': 1.0, 'name': 'S1', 'bus': 0}],
        'PV': [{'idx': 'G2', 'u': 1.0, 'name': 'G2', 'bus': 1}],
        'EDTSlot': [
            {'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1',
             'sd': '0.5,0.7', 'ug': '1,1'},
            {'idx': 'EDT2', 'u': 1.0, 'name': 'EDT2',
             'sd': '0.6,0.8', 'ug': '1,0'},
        ],
        'UCTSlot': [
            {'idx': 'UCT1', 'u': 1.0, 'name': 'UCT1', 'sd': '0.5,0.7'},
        ],
    }


class TestMigrate(unittest.TestCase):

    def test_legacy_csv_split(self):
        """Legacy CSV-list cells split into the right per-axis tables."""
        out, summary = migrate_data(_legacy_case())
        self.assertEqual(summary['ed_load'], 4)   # 2 areas * 2 slots
        self.assertEqual(summary['ed_gen'], 4)    # 2 gens  * 2 slots
        self.assertEqual(summary['uc_load'], 2)   # 2 areas * 1 slot
        self.assertEqual(len(out['EDSlotLoad']), 4)
        # Row content: (area, slot, sd) preserved with file ordering.
        first = out['EDSlotLoad'][0]
        self.assertEqual((first['area'], first['slot'], first['sd']),
                         (1, 'EDT1', 0.5))

    def test_top_level_key_rename(self):
        """``EDTSlot`` / ``UCTSlot`` keys rename to ``EDSlot`` / ``UCSlot``."""
        out, summary = migrate_data(_legacy_case())
        self.assertIn('EDSlot', out)
        self.assertIn('UCSlot', out)
        self.assertNotIn('EDTSlot', out)
        self.assertNotIn('UCTSlot', out)
        self.assertTrue(summary['renamed'])

    def test_idempotent_on_migrated_input(self):
        """Re-running on already-migrated data is a no-op."""
        out1, _ = migrate_data(_legacy_case())
        out2, summary2 = migrate_data(deepcopy(out1))
        self.assertTrue(summary2['no_op'])
        self.assertEqual(out1, out2)

    def test_partial_state_raises(self):
        """Mixed legacy cells + pre-existing per-axis table raises
        rather than silently dropping rows."""
        data = _legacy_case()
        # Manually pre-populate EDSlotLoad with one row to simulate
        # a partly-migrated case file.
        data['EDSlotLoad'] = [{'area': 1, 'slot': 'EDT1', 'sd': 9.99}]
        # Rename the legacy EDTSlot key to its new form too — only the
        # cells, not the table, were "migrated".
        with pytest.raises(ValueError, match="partly-migrated"):
            migrate_data(data)

    def test_csv_length_mismatch_raises(self):
        """``sd`` cell length not matching ``len(Area)`` raises with
        the slot idx + observed/expected lengths."""
        data = _legacy_case()
        # Length 3, areas=2 — and still comma-bearing so it's
        # detected as a legacy cell (a non-comma string would skip).
        data['EDTSlot'][0]['sd'] = '0.5,0.7,0.9'
        with pytest.raises(ValueError, match="EDT1.*length 3.*expected 2"):
            migrate_data(data)

    def test_no_area_drops_legacy_sd(self):
        """Cases without an Area block (npcc/wecc-style) drop their
        legacy sd cells with a warning instead of erroring."""
        data = {
            'Slack': [{'idx': 'S1', 'u': 1.0, 'name': 'S1', 'bus': 0}],
            'EDTSlot': [
                {'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1',
                 'sd': '0.5,0.7', 'ug': '1'},
            ],
        }
        out, summary = migrate_data(data)
        self.assertEqual(summary['ed_load'], 0)
        self.assertNotIn('EDSlotLoad', out)
        # EDSlot rows themselves are still slimmed.
        self.assertNotIn('sd', out['EDSlot'][0])

    def test_rename_only_pass(self):
        """A case already in v1.3.0 shape with only the legacy key
        names (no CSV cells) gets renamed, no per-axis emission."""
        data = {
            'Area': [{'idx': 1, 'u': 1.0, 'name': '1'}],
            'EDTSlot': [{'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1'}],
            'EDSlotLoad': [{'area': 1, 'slot': 'EDT1', 'sd': 0.5}],
        }
        out, summary = migrate_data(data)
        self.assertTrue(summary['renamed'])
        self.assertFalse(summary['no_op'])
        self.assertEqual(summary['ed_load'], 0)
        self.assertIn('EDSlot', out)
        self.assertEqual(len(out['EDSlotLoad']), 1)


class TestDuplicateSlotGuard(unittest.TestCase):

    def test_duplicate_slot_idx_raises(self):
        """Two ``EDTSlot`` rows with the same idx fail at file
        context, not silently downstream at solve time."""
        data = _legacy_case()
        data['EDTSlot'].append({'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1',
                                'sd': '0.5,0.7', 'ug': '1,1'})
        with pytest.raises(ValueError, match="duplicate slot idx 'EDT1'"):
            migrate_data(data)


class TestSingleAreaCell(unittest.TestCase):

    def test_single_area_no_comma_still_migrated(self):
        """A 1-area case has bare-numeric ``sd`` cells with no
        comma. The renamed-table path still recognizes them as
        legacy."""
        data = {
            'Area': [{'idx': 1, 'u': 1.0, 'name': '1'}],
            'EDTSlot': [
                {'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1', 'sd': '0.5'},
            ],
        }
        out, summary = migrate_data(data)
        self.assertEqual(summary['ed_load'], 1)
        self.assertEqual(out['EDSlotLoad'][0],
                         {'area': 1, 'slot': 'EDT1', 'sd': 0.5})
        # And the renamed definition row is slimmed.
        self.assertNotIn('sd', out['EDSlot'][0])


class TestEmptyStaticGenWithUg(unittest.TestCase):

    def test_no_staticgen_with_ug_data_raises(self):
        """``ug`` data on a case with no Slack/PV is a modeling bug."""
        data = {
            'Area': [{'idx': 1, 'u': 1.0, 'name': '1'}],
            'EDTSlot': [
                {'idx': 'EDT1', 'u': 1.0, 'name': 'EDT1',
                 'sd': '0.5', 'ug': '1,1,1'},
            ],
        }
        with pytest.raises(ValueError, match="no StaticGen entries"):
            migrate_data(data)


class TestJsonAdapter(unittest.TestCase):

    def test_disk_round_trip_idempotent(self):
        """Run the JSON adapter against an on-disk file twice; second
        invocation is a no-op."""
        data = _legacy_case()
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'case.json'
            src.write_text(json.dumps(data, indent=2) + '\n')

            # First pass: legacy split + rename.
            summary_a = migrate_json(src, src)
            self.assertFalse(summary_a.get('no_op', False))
            self.assertEqual(summary_a['ed_load'], 4)

            # Capture intermediate state.
            after_a = src.read_text()

            # Second pass: should be a no-op.
            summary_b = migrate_json(src, src)
            self.assertTrue(summary_b.get('no_op', False))
            self.assertEqual(src.read_text(), after_a)


class TestXlsxAdapter(unittest.TestCase):

    def test_xlsx_round_trip_renames_sheets_and_emits_tables(self):
        """End-to-end XLSX adapter: legacy sheets → migrated workbook
        with the new sheet names and per-axis tables, re-readable by
        openpyxl."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'case.xlsx'
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            # Area sheet.
            area_ws = wb.create_sheet('Area')
            area_ws.append(['idx', 'u', 'name'])
            area_ws.append([1, 1.0, '1'])
            area_ws.append([2, 1.0, '2'])
            # Slack sheet (one StaticGen).
            slack_ws = wb.create_sheet('Slack')
            slack_ws.append(['idx', 'u', 'name', 'bus'])
            slack_ws.append(['S1', 1.0, 'S1', 0])
            # Legacy EDTSlot + UCTSlot sheets with CSV cells.
            edt_ws = wb.create_sheet('EDTSlot')
            edt_ws.append(['idx', 'u', 'name', 'sd', 'ug'])
            edt_ws.append(['EDT1', 1.0, 'EDT1', '0.5,0.7', '1'])
            uct_ws = wb.create_sheet('UCTSlot')
            uct_ws.append(['idx', 'u', 'name', 'sd'])
            uct_ws.append(['UCT1', 1.0, 'UCT1', '0.5,0.7'])
            wb.save(str(src))

            summary = migrate_xlsx(src, src)
            self.assertEqual(summary['ed_load'], 2)
            self.assertEqual(summary['ed_gen'], 1)
            self.assertEqual(summary['uc_load'], 2)
            self.assertTrue(summary['renamed'])

            # Re-load and verify sheet names + content. ``read_only=True``
            # holds an open file handle on Windows, so close explicitly
            # before TemporaryDirectory cleanup tries to unlink the file.
            wb2 = openpyxl.load_workbook(str(src), read_only=True)
            try:
                self.assertIn('EDSlot', wb2.sheetnames)
                self.assertIn('UCSlot', wb2.sheetnames)
                self.assertIn('EDSlotLoad', wb2.sheetnames)
                self.assertIn('EDSlotGen', wb2.sheetnames)
                self.assertIn('UCSlotLoad', wb2.sheetnames)
                self.assertNotIn('EDTSlot', wb2.sheetnames)
                self.assertNotIn('UCTSlot', wb2.sheetnames)

                edload_rows = list(wb2['EDSlotLoad'].iter_rows(values_only=True))
                self.assertEqual(edload_rows[0], ('area', 'slot', 'sd'))
                # 2 areas * 1 slot = 2 data rows.
                self.assertEqual(len(edload_rows), 3)
            finally:
                wb2.close()


class TestRenameLegacyKeys(unittest.TestCase):

    def test_keys_renamed_in_order(self):
        data = {'A': 1, 'EDTSlot': [], 'B': 2, 'UCTSlot': []}
        out, renamed = _rename_legacy_keys(data)
        self.assertTrue(renamed)
        self.assertEqual(list(out.keys()), ['A', 'EDSlot', 'B', 'UCSlot'])

    def test_no_rename_when_absent(self):
        data = {'A': 1, 'EDSlot': []}
        out, renamed = _rename_legacy_keys(data)
        self.assertFalse(renamed)
        self.assertEqual(out, data)
